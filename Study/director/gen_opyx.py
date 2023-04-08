import os
import re
import time
import openpyxl
from openpyxl import load_workbook
import datetime
from openpyxl.chart import LineChart,Reference,BarChart,PieChart
from openpyxl.chart.series import DataPoint

def  read_tx(read_path):
    '从文件中提取字符并返回列表'
    aa = []
    dicdt = {}
    dicdt['error'] = 0
    with open(read_path,'r',encoding='utf-8') as f:
        for line in f.readlines():
            line = line[:-1]
            words = line.split()
            for word in words:
                aa.append(word)
        print(aa)

        # for line in f:
        #     line = line[:-1]
        #     words = line.split()
        #     print(words)
        #     for word in words:
        #         if '88' != word:
        #             continue
        #         else:
        #             dicdt['error'] += 1
        # print(dicdt)
    # return aa
def write_opyx_excel(save_path,datas):
    '利用openpyxl第三方库将数据写入表格'
    day_time = time.strftime('%Y_%m_%d',time.localtime())
    wb = openpyxl.Workbook()   #创建工作对象
    sheet = wb.create_sheet('类型比较',0)  #创建附表并按照索引确认位置
    # sheet.cell(1, 1).value = '类型'
    # sheet.cell(1, 2).value = '数量'
    wb.save(save_path)
    for col in range(len(datas)):
        for row in range(len(datas[col])):
            sheet.cell(row + 1, col + 1).value = datas[col][row]
    wb.save(save_path)
    if not os.path.exists(save_path):
        print(f'{save_path}不存在')
    else:
        print('{}文件保存成功'.format(save_path))
    return day_time

def gen_line_opxy(save_path):
    '利用openpyxl第三方库表格中的数据生成折线图看板'
    wb = openpyxl.load_workbook(save_path)   #load_workbook读excel表格
    sheet = wb['类型比较']
    # sheet = wb.active
    line = LineChart()
    line.title='折线图'
    data = Reference(sheet,min_row=2,max_row=6,min_col=1,max_col=3)
    title_data = Reference(sheet,min_row=1,max_row=1,min_col=2,max_col=3)
    line.add_data(data,from_rows=True,titles_from_data=True)
    line.set_categories(title_data)
    sheet.add_chart(line,'H10')   #将看板放置在附表的规定位置上
    wb.save(save_path)
def gen_bar_opxy(save_path):
    '利用openpyxl第三方库表格中的数据生成柱状图看板'
    wb = openpyxl.load_workbook(save_path)
    sheet = wb['类型比较']
    bar = BarChart()
    bar.title='柱状图'
    data = Reference(sheet, min_row=2, max_row=6, min_col=1, max_col=3)
    title_data = Reference(sheet, min_row=1, max_row=1, min_col=2, max_col=3)
    bar.add_data(data, from_rows=True, titles_from_data=True)
    bar.set_categories(title_data)
    sheet.add_chart(bar, 'H26')
    wb.save(save_path)
# def gen_pie_opxy(save_path):
#    '利用openpyxl第三方库表格中的数据生成饼图看板'
#     wb = openpyxl.load_workbook(save_path)
#     sheet = wb['2023_04_05类型比较']
#     pie = PieChart()
#     pie.title = '饼图'
#     data = Reference(sheet,  min_col=2,min_row=1, max_row=6)
#     title_data = Reference(sheet, min_col=1,min_row=2, max_row=6)
#     pie.add_data(data, from_rows=True, titles_from_data=True)
#     pie.set_categories(title_data)
#     silce = DataPoint(idx=0,explosion=20)
#     pie.series[0].data_points = [silce]
#     sheet.add_chart(pie, 'H41')
#     wb.save(save_path)

if __name__ == '__main__':
    list1 = ['error','类型1', '类型2', '类型3', '类型4', '类型5']
    list2 = ['数量1',22, 66, 55, 44, 88]
    list3 = ['数量2',15, 15, 63, 44, 22]
    lists = []
    lists.append(list1)
    lists.append(list2)
    lists.append(list3)
    # print(lists)
    # read_path = 'D:/study/yy.txt'
    # read_tx(read_path)
    save_path = input('请输入你想要保存的路径及表格名称：')
    write_opyx_excel(save_path,lists)
    gen_line_opxy(save_path)
    gen_bar_opxy(save_path)
    # gen_pie_opxy(save_path)


