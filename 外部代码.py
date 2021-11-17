from openpyxl import load_workbook
from InterfaceAutoTest.utils.read_json import read_json
from InterfaceAutoTest.data.excel_column import ExcelColumn
from InterfaceAutoTest.utils.read_ini import ReadIni

class ReadExcel:
    def __init__(self):
        self.excel_path = ReadIni().get_excel_path()
        self.data_json_path=ReadIni().get_data_json_path()
        self.expect_json_path = ReadIni().get_expect_json_path()
        wb = load_workbook(self.excel_path)
        # 读取sheet页
        self.ws = wb["接口自动化用例模板"]

    """
    分别声明方法来实现Excel每一列数据的读取，比如get_id来读取用例id，get_name来读取
    用例name...
    """
    # 声明一个方法，读取用例id
    def get_case_id(self, row):
        return self.ws[ExcelColumn.CASE_ID + str(row)].value

    # 声明一个方法，读取用例模块
    def get_case_module(self, row):
        return self.ws[ExcelColumn.CASE_MODULE + str(row)].value

    # 声明一个方法，读取用例名称
    def get_case_name(self, row):
        return self.ws[ExcelColumn.CASE_NAME + str(row)].value

    # 读取用例是够执行
    def get_if_run(self,row):
        return self.ws[ExcelColumn.CASE_IF_RUN+str(row)].value

    # 读取用例请求方法
    def get_case_method(self,row):
        return self.ws[ExcelColumn.CASE_METHOD+str(row)].value

    # 读取用例url
    def get_case_url(self, row):
        return self.ws[ExcelColumn.CASE_URL + str(row)].value

    # 读取用例数据
    def get_case_data(self, row):
        key= self.ws[ExcelColumn.CASE_DATA + str(row)].value
        #如果key不为None，说明这个接口需要传参，这时再去掉读JSON文件中的数据
        if key:
            return read_json(self.data_json_path)[key]

    # 读取断言数据
    def get_case_expect(self,row):
        key =  self.ws[ExcelColumn.CASE_EXPECT+str(row)].value
        if key:
            return read_json(self.expect_json_path)[key]

    #获取Excel的行数
    def get_row_count(self):
        return self.ws.max_row

    #获取Excel的列数
    def get_column_count(self):
        return self.ws.max_column

    #获取前置用例id
    def get_precondition_id(self,row):
        return self.ws[ExcelColumn.CASE_PRECONDITION_ID + str(row)].value

    #获取依赖的字段key
    def get_depend_key(self,row):